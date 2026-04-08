"""Download actual distances from Garmin Connect and write them into the XLS."""

from __future__ import annotations

import datetime
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook

import xlrd

from .garmin_client import GarminClient
from .parser import WorkoutRow

# "km í raun" lives in column 4 (0-indexed in xlrd, 1-indexed in openpyxl → col 5)
_ACTUAL_KM_COL_OPENPYXL = 5  # openpyxl is 1-indexed


def round_to_increment(value: float, increment: float) -> float:
    """Round value to the nearest multiple of increment.

    >>> round_to_increment(9.73, 0.5)
    9.5
    >>> round_to_increment(9.75, 0.5)
    10.0
    >>> round_to_increment(10.26, 0.5)
    10.5
    """
    return round(value / increment) * increment


def _build_date_distance_map(
    activities: list[dict],
) -> dict[datetime.date, float]:
    """Sum running distances per date from Garmin activity list.

    Activities have distance in meters and startTimeLocal as ISO string.
    """
    totals: dict[datetime.date, float] = defaultdict(float)
    for activity in activities:
        distance_m = activity.get("distance", 0) or 0
        start_local = activity.get("startTimeLocal", "")
        if not start_local or distance_m <= 0:
            continue
        date = datetime.date.fromisoformat(start_local[:10])
        totals[date] += distance_m / 1000.0
    return dict(totals)


def fetch_actual_distances(
    client: GarminClient,
    rows: list[WorkoutRow],
) -> dict[datetime.date, float]:
    if not rows:
        return {}

    start = min(r.date for r in rows)
    end = max(r.date for r in rows)
    activities = client.get_activities_by_date(start.isoformat(), end.isoformat())
    return _build_date_distance_map(activities)


def _convert_xls_to_xlsx(xls_path: Path) -> Workbook:
    rd = xlrd.open_workbook(str(xls_path))
    sheet = rd.sheet_by_index(0)

    wb = Workbook()
    ws = wb.active
    assert ws is not None

    for row_idx in range(sheet.nrows):
        for col_idx in range(sheet.ncols):
            cell_type = sheet.cell_type(row_idx, col_idx)
            value = sheet.cell_value(row_idx, col_idx)

            if cell_type == xlrd.XL_CELL_DATE:
                date_tuple = xlrd.xldate_as_tuple(float(value), rd.datemode)
                value = datetime.datetime(*date_tuple)
            elif cell_type == xlrd.XL_CELL_EMPTY:
                continue

            ws.cell(row=row_idx + 1, column=col_idx + 1, value=value)

    return wb


def write_actual_km(
    xls_path: Path,
    rows: list[WorkoutRow],
    distances: dict[datetime.date, float],
    increment: float,
) -> list[tuple[WorkoutRow, float]]:
    """Write rounded distances into column 4. Saves as .xlsx (openpyxl only writes xlsx)."""
    xlsx_path = xls_path.with_suffix(".xlsx")

    if xlsx_path.exists():
        wb = load_workbook(str(xlsx_path))
    else:
        wb = _convert_xls_to_xlsx(xls_path)

    ws = wb.active
    if ws is None:
        raise ValueError(f"No active sheet in {xls_path}")

    written: list[tuple[WorkoutRow, float]] = []

    for row in rows:
        raw_km = distances.get(row.date)
        if raw_km is None:
            continue

        rounded = round_to_increment(raw_km, increment)
        # openpyxl rows are 1-indexed; xlrd row_idx is 0-indexed → +1
        ws.cell(row=row.row_idx + 1, column=_ACTUAL_KM_COL_OPENPYXL, value=rounded)
        written.append((row, rounded))

    wb.save(str(xlsx_path))
    return written
