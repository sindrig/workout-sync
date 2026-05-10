"""Parser module for coach's workout training plan XLS/XLSX files (Icelandic format)."""

from __future__ import annotations

import datetime
from dataclasses import dataclass
from pathlib import Path

import xlrd
from openpyxl import load_workbook


@dataclass
class Workout:
    """A single workout entry from the training plan."""

    date: datetime.date
    day_name: str
    workout_type: str
    description: str
    distance_km: float
    notes: str


@dataclass
class WorkoutRow:
    """A date row in the XLS, including its row index for write-back."""

    row_idx: int
    date: datetime.date
    day_name: str
    description: str
    distance_km: float


class _XlrdSheet:
    def __init__(self, filepath: str):
        self._wb = xlrd.open_workbook(filepath)
        self._sheet = self._wb.sheet_by_index(0)

    @property
    def nrows(self) -> int:
        return self._sheet.nrows

    def is_date(self, row: int, col: int) -> bool:
        return self._sheet.cell_type(row, col) == xlrd.XL_CELL_DATE

    def date_value(self, row: int, col: int) -> datetime.date:
        val = float(self._sheet.cell_value(row, col))
        t = xlrd.xldate_as_tuple(val, self._wb.datemode)
        return datetime.date(*t[:3])

    def string_value(self, row: int, col: int) -> str:
        return str(self._sheet.cell_value(row, col)).strip()

    def numeric_value(self, row: int, col: int) -> float | None:
        if self._sheet.cell_type(row, col) == xlrd.XL_CELL_NUMBER:
            return float(self._sheet.cell_value(row, col))
        return None


class _OpenpyxlSheet:
    def __init__(self, filepath: str):
        self._wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = self._wb.active
        if ws is None:
            raise ValueError("XLSX file has no active worksheet")
        self._sheet = ws
        self._rows = list(self._sheet.iter_rows())

    @property
    def nrows(self) -> int:
        return len(self._rows)

    def is_date(self, row: int, col: int) -> bool:
        cell = self._rows[row][col]
        return isinstance(cell.value, datetime.datetime)

    def date_value(self, row: int, col: int) -> datetime.date:
        val = self._rows[row][col].value
        if isinstance(val, datetime.datetime):
            return val.date()
        if isinstance(val, datetime.date):
            return val
        raise TypeError(f"Expected date at ({row}, {col}), got {type(val)}")

    def string_value(self, row: int, col: int) -> str:
        val = self._rows[row][col].value
        return str(val).strip() if val is not None else ""

    def numeric_value(self, row: int, col: int) -> float | None:
        val = self._rows[row][col].value
        if isinstance(val, (int, float)):
            return float(val)
        return None


def _open_sheet(filepath: str) -> _XlrdSheet | _OpenpyxlSheet:
    if Path(filepath).suffix.lower() == ".xlsx":
        return _OpenpyxlSheet(filepath)
    return _XlrdSheet(filepath)


def _classify_workout_type(description: str) -> str:
    """Classify workout type from Icelandic description text.

    Order matters — check more specific terms before general ones.
    """
    desc_lower = description.lower()

    if "ról" in desc_lower:
        return "ról"
    if "hraðaæf" in desc_lower:
        return "hraðaæf"
    if "jafnt" in desc_lower:
        return "jafnt"
    if "samæf" in desc_lower:
        return "samæfing"
    if "styrktaræfing" in desc_lower:
        return "styrktaræfing"
    if "fartleikur" in desc_lower or "fartleik" in desc_lower:
        return "fartleikur"
    if "hlaupasería" in desc_lower:
        return "samæfing"
    return "other"


def parse_xls(filepath: str) -> list[Workout]:
    """Parse coach's workout XLS/XLSX file and extract non-rest workout days.

    Args:
        filepath: Path to the XLS or XLSX file.

    Returns:
        List of Workout entries sorted by date ascending, with rest days excluded.
    """
    sheet = _open_sheet(filepath)
    workouts: list[Workout] = []

    for row_idx in range(8, sheet.nrows):
        if not sheet.is_date(row_idx, 0):
            continue

        date = sheet.date_value(row_idx, 0)
        day_name = sheet.string_value(row_idx, 1)
        description = sheet.string_value(row_idx, 2)
        distance_km = sheet.numeric_value(row_idx, 3) or 0.0
        notes = sheet.string_value(row_idx, 5)

        workout_type = _classify_workout_type(description)

        if distance_km == 0.0 and workout_type != "styrktaræfing":
            continue

        workouts.append(
            Workout(
                date=date,
                day_name=day_name,
                workout_type=workout_type,
                description=description,
                distance_km=distance_km,
                notes=notes,
            )
        )

    workouts.sort(key=lambda w: w.date)
    return workouts


def parse_xls_rows(filepath: str) -> list[WorkoutRow]:
    """Parse all date rows from the XLS/XLSX, unfiltered, with row indices for write-back."""
    sheet = _open_sheet(filepath)
    rows: list[WorkoutRow] = []

    for row_idx in range(8, sheet.nrows):
        if not sheet.is_date(row_idx, 0):
            continue

        date = sheet.date_value(row_idx, 0)
        day_name = sheet.string_value(row_idx, 1)
        description = sheet.string_value(row_idx, 2)
        distance_km = sheet.numeric_value(row_idx, 3) or 0.0

        rows.append(
            WorkoutRow(
                row_idx=row_idx,
                date=date,
                day_name=day_name,
                description=description,
                distance_km=distance_km,
            )
        )

    rows.sort(key=lambda r: r.date)
    return rows
